#!/usr/bin/env python3
"""Ingest a "Pull Requests and AI" xlsx export into per-sheet CSVs in
content/documents/uploads/.

Accepts either:
  - the JSON file the Google Drive MCP `download_file_content` tool saves when
    its result is too large for context ({content: <base64>, title, mimeType}), or
  - a raw .xlsx path.

Usage:
    python3 pipeline/ingest_email_export.py /path/to/tool-result.json
    python3 pipeline/ingest_email_export.py /path/to/export.xlsx

Writes one RFC-4180 CSV per sheet:
    content/documents/uploads/<title> - <Sheet>.csv
(e.g. "Pull Requests and AI 2026 07 14 - Pull Requests.csv"). These CSVs are
the canonical source for pipeline/refresh_copilot.py, which picks the newest
pair by filename.
"""

import base64
import csv
import json
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
UPLOADS_DIR = PROJECT_ROOT / "content" / "documents" / "uploads"
XLSX_MAGIC = b"PK\x03\x04"


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"ERROR: input not found: {src}")

    if src.suffix.lower() in (".xlsx", ".xls"):
        xlsx_path, title = src, src.stem
    else:
        d = json.loads(src.read_text(encoding="utf-8"))
        title = Path(d.get("title") or "export.xlsx").stem
        raw = base64.b64decode(d["content"])
        if raw[:4] != XLSX_MAGIC:
            sys.exit(f"ERROR: decoded content is not an xlsx (magic {raw[:4]!r})")
        xlsx_path = Path(tempfile.mkdtemp()) / f"{title}.xlsx"
        xlsx_path.write_bytes(raw)
        print(f"decoded {len(raw)} bytes -> {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    written = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
            rows.pop()
        out = UPLOADS_DIR / f"{title} - {ws.title}.csv"
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for r in rows:
                w.writerow(["" if c is None else c for c in r])
        print(f"  {ws.title!r}: {len(rows)} rows -> {out.relative_to(PROJECT_ROOT)}")
        written.append(out)
    if not written:
        sys.exit("ERROR: workbook has no sheets")


if __name__ == "__main__":
    main()
